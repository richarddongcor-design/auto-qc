import pytest
import tempfile
from pathlib import Path
import json
from auto_qc.qc.domain.report import write_report, verify_report_exists


def _make_sample_inputs():
    rule_name_map = {
        "auto-pi_R01": "答非所问",
        "auto-pi_R02": "过度追问",
        "project_R01": "未介绍产品",
    }
    wide_rows = [
        {
            "id": "001", "time": "06-01", "intent": "B",
            "rules": {
                "auto-pi_R01": {"result": "违规", "evidence": "用户问XX AI说YY", "rule_name": "答非所问"},
                "auto-pi_R02": {"result": "通过", "evidence": "", "rule_name": "过度追问"},
                "project_R01": {"result": "违规", "evidence": "AI未介绍产品", "rule_name": "未介绍产品"},
            },
            "summary": "答非所问：用户问XX AI说YY\n未介绍产品：AI未介绍产品",
        },
        {
            "id": "002", "time": "06-01", "intent": "A",
            "rules": {
                "auto-pi_R01": {"result": "通过", "evidence": "", "rule_name": "答非所问"},
                "auto-pi_R02": {"result": "通过", "evidence": "", "rule_name": "过度追问"},
                "project_R01": {"result": "通过", "evidence": "", "rule_name": "未介绍产品"},
            },
            "summary": "",
        },
    ]
    stats = {
        "total": 2,
        "pass_count": 1,
        "violation_count": 1,
        "violation_rate": "50.0%",
        "rule_set_stats": {
            "auto-pi": {"total_checks": 4, "violations": 1, "rate": "25.0%"},
            "project": {"total_checks": 2, "violations": 1, "rate": "50.0%"},
        },
        "rule_stats": {
            "auto-pi_R01": {"pass": 1, "violation": 1, "pass_rate": "50.0%", "violation_rate": "50.0%"},
            "auto-pi_R02": {"pass": 2, "violation": 0, "pass_rate": "100.0%", "violation_rate": "0.0%"},
            "project_R01": {"pass": 1, "violation": 1, "pass_rate": "50.0%", "violation_rate": "50.0%"},
        },
        "problem_distribution": [
            {"rule_id": "auto-pi_R01", "rule_name": "答非所问", "count": 1, "ratio": "100.0%"},
            {"rule_id": "project_R01", "rule_name": "未介绍产品", "count": 1, "ratio": "100.0%"},
        ],
        "rule_name_map": rule_name_map,
    }
    return wide_rows, stats


def test_write_report_creates_file():
    with tempfile.TemporaryDirectory() as tmp:
        output = str(Path(tmp) / "report.xlsx")
        wide_rows, stats = _make_sample_inputs()
        write_report(output, wide_rows, stats)
        assert verify_report_exists(output)


def test_report_has_two_sheets():
    import openpyxl
    with tempfile.TemporaryDirectory() as tmp:
        output = str(Path(tmp) / "report.xlsx")
        wide_rows, stats = _make_sample_inputs()
        write_report(output, wide_rows, stats)
        wb = openpyxl.load_workbook(output)
        assert "打标明细" in wb.sheetnames
        assert "统计概览" in wb.sheetnames


def test_report_wide_table_headers():
    import openpyxl
    with tempfile.TemporaryDirectory() as tmp:
        output = str(Path(tmp) / "report.xlsx")
        wide_rows, stats = _make_sample_inputs()
        write_report(output, wide_rows, stats)
        wb = openpyxl.load_workbook(output)
        ws = wb["打标明细"]
        assert ws.max_column == 7  # id | 时间 | 意向 | 3规则列 | 打标详情
        summary_cell = ws.cell(2, 7).value
        assert "答非所问" in summary_cell


def test_report_shows_violation_in_cell():
    import openpyxl
    with tempfile.TemporaryDirectory() as tmp:
        output = str(Path(tmp) / "report.xlsx")
        wide_rows, stats = _make_sample_inputs()
        write_report(output, wide_rows, stats)
        wb = openpyxl.load_workbook(output)
        ws = wb["打标明细"]
        # Row 2: 001 should have "违规" in column 4 (auto-pi_R01)
        assert ws.cell(2, 4).value == "违规"
        # Row 2: 001 should have "通过" in column 5 (auto-pi_R02)
        assert ws.cell(2, 5).value == "通过"
        # Row 3: 002 should have "通过" in column 4
        assert ws.cell(3, 4).value == "通过"
