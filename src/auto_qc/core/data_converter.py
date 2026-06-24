"""统一数据转换 — 模板生成 + xlsx → JSON + 智能列检测。

功能：
1. 生成 xlsx 模板（用户填写后上传）
2. 智能检测对话列：JSON → 自动转可读文本；纯文本 → 直接返回
3. 输出统一格式 JSON，QC 和 PI 共用
"""
import json
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


# ─── 常量 ───

HEADERS = ["ID", "对话内容"]
HEADER_FILL = PatternFill(start_color="F5F0E8", end_color="F5F0E8", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11)
EXAMPLE_ID = "001001"
EXAMPLE_CONV = '[{"ttsResult":"您好，这边是XX招聘","asrResult":"你好，请问是哪家公司"}]'


# ─── 模板生成 ───


def generate_template(path: str | Path) -> str:
    """生成 xlsx 模板文件，返回文件路径。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "对话数据"

    # 表头
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # 示例行
    ws.cell(row=2, column=1, value=EXAMPLE_ID)
    ws.cell(row=2, column=2, value=EXAMPLE_CONV)

    # 列宽
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 60

    wb.save(str(path))
    return str(path)


# ─── 智能对话列检测 ───


def _is_json(text: str) -> bool:
    """检测文本是否为合法 JSON。"""
    t = text.strip()
    if not t:
        return False
    if t[0] not in ("[", "{"):
        return False
    try:
        json.loads(t)
        return True
    except json.JSONDecodeError:
        return False


def _is_plain_conversation(text: str) -> bool:
    """检测文本是否已是可读对话格式（含 AI: 或 用户: 标记）。"""
    return "AI:" in text or "用户:" in text or "客服:" in text or "客户:" in text


def smart_preprocess(conv_raw: str) -> str:
    """智能预处理对话列。

    检测策略：
    1. JSON 格式 → 解析并转为 `AI: xxx\n用户: xxx` 可读文本
    2. 已是纯文本格式 → 直接返回
    3. 其他 → 原样返回
    """
    text = str(conv_raw).strip()
    if not text:
        return ""

    # 检测是否为 JSON
    if _is_json(text):
        # 双重编码处理（Excel 可能把 JSON 再包一层引号）
        if text.startswith('"[') or text.startswith("\"["):
            try:
                text = json.loads(text)
            except (json.JSONDecodeError, TypeError):
                pass

        # 解析 JSON
        try:
            if isinstance(text, str):
                data = json.loads(text)
            else:
                data = text

            # 转换成可读文本
            lines = []
            for turn in data if isinstance(data, list) else [data]:
                tts = (turn.get("ttsResult") or turn.get("tts", "")).strip()
                asr = (turn.get("asrResult") or turn.get("asr", "")).strip()
                if tts:
                    lines.append(f"AI: {tts}")
                if asr:
                    lines.append(f"用户: {asr}")
            if lines:
                return "\n".join(lines)
        except (json.JSONDecodeError, TypeError, AttributeError):
            pass

    # 已是纯文本对话 → 直接返回
    if _is_plain_conversation(text):
        return text

    # 兜底：原样返回
    return text


def convert_xlsx(path: str | Path, id_col: str = "ID", conv_col: str = "对话内容") -> list[dict]:
    """读取 xlsx，智能转换对话列，返回统一 JSON。

    Returns:
        [{"id": str, "conversation": str}, ...]
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    headers = [str(h).strip() if h else "" for h in next(rows)]

    # 匹配列
    def _find_col(name_zh: str, alt: str) -> int | None:
        for i, h in enumerate(headers):
            if name_zh in h or alt.lower() in h.lower():
                return i
        return None

    id_idx = _find_col("ID", "id")
    conv_idx = _find_col("对话内容", "conversation")

    if id_idx is None:
        raise ValueError("未找到 ID 列")
    if conv_idx is None:
        raise ValueError("未找到「对话内容」列")

    results = []
    for row in rows:
        if row[id_idx] is None:
            continue
        raw_conv = str(row[conv_idx]) if row[conv_idx] and conv_idx is not None else ""
        results.append({
            "id": str(row[id_idx]).strip(),
            "conversation": smart_preprocess(raw_conv),
        })

    wb.close()

    if not results:
        raise ValueError("未读取到任何有效数据")

    return results
