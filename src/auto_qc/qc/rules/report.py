"""质检报告 Excel 生成（v2.0 宽表模式）"""
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
PASS_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
VIOLATION_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")


def _style_header(cell):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_report(output_path: str, wide_rows: list[dict], stats: dict) -> None:
    """生成宽表质检报告 Excel。"""
    wb = openpyxl.Workbook()

    _write_detail_sheet(wb, wide_rows)
    _write_stats_sheet(wb, stats)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()


def _write_detail_sheet(wb, wide_rows: list[dict]) -> None:
    """Sheet 1: 打标明细（宽表）。"""
    ws = wb.active
    ws.title = "打标明细"
    ws.sheet_properties.tabColor = "4472C4"

    if not wide_rows:
        ws.cell(1, 1, "无数据")
        return

    rule_ids = list(wide_rows[0]["rules"].keys())
    rule_names = {rid: wide_rows[0]["rules"][rid].get("rule_name", rid) for rid in rule_ids}

    headers = ["id", "时间"] + [f"{rid}: {rule_names[rid]}" for rid in rule_ids] + ["打标详情"]
    for col, h in enumerate(headers, 1):
        _style_header(ws.cell(1, col, h))

    for row_idx, row in enumerate(wide_rows, 2):
        ws.cell(row_idx, 1, row["id"])
        ws.cell(row_idx, 2, row.get("time", ""))

        for col_idx, rid in enumerate(rule_ids, 3):
            rr = row["rules"].get(rid, {})
            result = rr.get("result", "通过")
            cell = ws.cell(row_idx, col_idx, result)
            cell.fill = VIOLATION_FILL if result == "违规" else PASS_FILL

        summary_col = 3 + len(rule_ids)
        ws.cell(row_idx, summary_col, row.get("summary", ""))

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 16
    for col_idx in range(3, 3 + len(rule_ids)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14
    ws.column_dimensions[get_column_letter(3 + len(rule_ids))].width = 60


def _write_stats_sheet(wb, stats: dict) -> None:
    """Sheet 2: 统计概览。"""
    ws = wb.create_sheet("统计概览")
    ws.sheet_properties.tabColor = "FFD93D"

    row = 1

    # 总体概览
    ws.cell(row, 1, "【总体概览】").font = Font(bold=True, size=13)
    row += 1
    for label, key in [("总对话数", "total"), ("违规对话数", "violation_count"),
                        ("通过对话数", "pass_count"), ("总体违规率", "violation_rate")]:
        ws.cell(row, 1, label)
        ws.cell(row, 2, stats.get(key, 0))
        row += 1
    row += 1

    # 按规则集统计
    ws.cell(row, 1, "【按规则集统计】").font = Font(bold=True, size=13)
    row += 1
    for col, h in enumerate(["规则集", "总检查次数", "违规次数", "违规率"], 1):
        _style_header(ws.cell(row, col, h))
    row += 1
    for rs_name, rs_stat in stats.get("rule_set_stats", {}).items():
        ws.cell(row, 1, rs_name)
        ws.cell(row, 2, rs_stat.get("total_checks", 0))
        ws.cell(row, 3, rs_stat.get("violations", 0))
        ws.cell(row, 4, rs_stat.get("rate", "0%"))
        row += 1
    row += 1

    # 按规则统计
    ws.cell(row, 1, "【按规则统计】").font = Font(bold=True, size=13)
    row += 1
    for col, h in enumerate(["规则ID", "规则名称", "规则集", "通过数", "违规数", "通过率", "违规率"], 1):
        _style_header(ws.cell(row, col, h))
    row += 1
    rule_name_map = stats.get("rule_name_map", {})
    for rid in sorted(stats.get("rule_stats", {}).keys()):
        s = stats["rule_stats"][rid]
        rs_name = rid.split("_")[0] if "_" in rid else ""
        ws.cell(row, 1, rid)
        ws.cell(row, 2, rule_name_map.get(rid, rid))
        ws.cell(row, 3, rs_name)
        ws.cell(row, 4, s["pass"])
        ws.cell(row, 5, s["violation"])
        ws.cell(row, 6, s.get("pass_rate", ""))
        ws.cell(row, 7, s.get("violation_rate", ""))
        row += 1
    row += 1

    # 问题分布
    ws.cell(row, 1, "【有违规case中的问题分布】").font = Font(bold=True, size=13)
    row += 1
    ws.cell(row, 1, f"（在 {stats.get('violation_count', 0)} 个至少有一条违规的对话中）")
    row += 1
    for col, h in enumerate(["问题类型", "出现次数", "占违规case比例"], 1):
        _style_header(ws.cell(row, col, h))
    row += 1
    for pd in stats.get("problem_distribution", []):
        ws.cell(row, 1, f"{pd['rule_id']}: {pd['rule_name']}")
        ws.cell(row, 2, pd["count"])
        ws.cell(row, 3, pd["ratio"])
        row += 1

    ws.column_dimensions["A"].width = 30
    for c in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[c].width = 12


def verify_report_exists(output_path: str) -> bool:
    """验证报告文件是否生成且非空。"""
    p = Path(output_path)
    return p.exists() and p.stat().st_size > 0
