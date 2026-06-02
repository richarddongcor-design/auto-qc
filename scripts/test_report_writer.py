"""测试 report_writer.py 的报告生成和清理功能"""
import json
import sys
import tempfile
from pathlib import Path

import pytest

# Make sure we can import from the same directory
sys.path.insert(0, str(Path(__file__).parent))

from report_writer import write_report, cleanup_temp_files


def test_write_report_basic(tmp_path):
    """测试基本报告生成"""
    qc_results = [
        {"id": "1", "time": "2026-06-01", "intent": "A(有意向)", "violations": []},
        {
            "id": "2",
            "time": "2026-06-01",
            "intent": "B(不确定)",
            "violations": [
                {
                    "rule_id": "R01",
                    "rule_name": "无视用户明确拒绝",
                    "severity": "高",
                    "evidence": "用户说不考虑，AI继续...",
                    "suggestion": "应礼貌结束",
                }
            ],
        },
    ]
    attribution_results = {
        "B(不确定)": [
            {
                "category": "未介绍岗位亮点",
                "count": 5,
                "ratio": 0.20,
                "examples": ["对话 123..."],
                "suggestion": "应主动介绍薪资地点",
            }
        ]
    }
    stats = {"total": 2, "violation_rate": 0.5, "rules_hit": {"R01": 1}}

    output_path = str(tmp_path / "report.xlsx")
    write_report(output_path, qc_results, attribution_results, stats)

    assert Path(output_path).exists()

    # 验证 sheet 存在
    import openpyxl
    wb = openpyxl.load_workbook(output_path)
    assert "合规检测" in wb.sheetnames
    assert "归因分析" in wb.sheetnames
    assert "统计概览" in wb.sheetnames
    wb.close()


def test_cleanup_temp_files(tmp_path):
    """测试临时文件清理"""
    (tmp_path / "progress.json").write_text("{}")
    (tmp_path / "failed_batches.json").write_text("{}")
    (tmp_path / "batches" / "batch_1.json").parent.mkdir(parents=True, exist_ok=True)
    (tmp_path / "batches" / "batch_1.json").write_text("{}")

    cleanup_temp_files(str(tmp_path))

    assert not (tmp_path / "progress.json").exists()
    assert not (tmp_path / "failed_batches.json").exists()
    assert not (tmp_path / "batches").exists()


def test_cleanup_keep_temp(tmp_path):
    """测试保留临时文件"""
    (tmp_path / "progress.json").write_text("{}")
    cleanup_temp_files(str(tmp_path), keep_temp=True)
    assert (tmp_path / "progress.json").exists()


def test_stats_table_format(tmp_path):
    """测试统计概览以规则明细表格式展示（非 JSON 字符串）"""
    qc_results = [
        {"id": "1", "time": "2026-06-01", "intent": "A(有意向)", "violations": []},
    ]
    attribution_results = {}
    stats = {
        "total": 1,
        "pass": 1,
        "violation_rate": "0.0%",
        "rules_hit": {"R01": 5, "R02": 10}
    }

    output_path = str(tmp_path / "report.xlsx")
    write_report(output_path, qc_results, attribution_results, stats)

    import openpyxl
    wb = openpyxl.load_workbook(output_path)
    ws = wb["统计概览"]

    # 验证基础指标行
    assert ws.cell(1, 1).value == "指标"
    assert ws.cell(2, 1).value == "总对话数"
    assert ws.cell(2, 2).value == 1
    assert ws.cell(4, 1).value == "违规率"

    # 验证规则明细表表头
    assert ws.cell(6, 1).value == "规则ID"
    assert ws.cell(6, 2).value == "规则名称"
    assert ws.cell(6, 3).value == "命中次数"
    assert ws.cell(6, 4).value == "占比"

    # 验证规则数据行
    assert ws.cell(7, 1).value == "R01"
    assert ws.cell(7, 3).value == 5
    assert ws.cell(7, 4).value == "33.3%"

    assert ws.cell(8, 1).value == "R02"
    assert ws.cell(8, 3).value == 10
    assert ws.cell(8, 4).value == "66.7%"

    # 验证合计行
    assert ws.cell(9, 1).value == "合计"
    assert ws.cell(9, 3).value == 15

    wb.close()


def test_cleanup_attribution_batches(tmp_path):
    """测试清理归因批次目录"""
    (tmp_path / "batches" / "batch_1.json").parent.mkdir(parents=True, exist_ok=True)
    (tmp_path / "batches" / "batch_1.json").write_text("{}")
    (tmp_path / "attribution_batches" / "batch_1.json").parent.mkdir(parents=True, exist_ok=True)
    (tmp_path / "attribution_batches" / "batch_1.json").write_text("{}")

    cleanup_temp_files(str(tmp_path))

    assert not (tmp_path / "batches").exists()
    assert not (tmp_path / "attribution_batches").exists()
