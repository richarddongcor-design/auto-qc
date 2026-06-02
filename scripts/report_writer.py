"""
report_writer.py — 质检报告 Excel 生成 + 临时文件清理

子命令：
  python report_writer.py write --output <报告路径> --qc-results <JSON路径> --attribution <JSON路径>
  python report_writer.py cleanup --dir <目录路径> [--keep-temp]
"""

import argparse
import json
import os
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


# ─── 样式常量 ───

HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SEVERITY_FILLS = {
    "高": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
    "中": PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid"),
    "低": PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid"),
}
NORMAL_ALIGNMENT = Alignment(wrap_text=True, vertical="top")


def _style_header(cell):
    """设置表头样式。"""
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_row(row, severity=None):
    """设置数据行样式，第 6 列（危害程度）按严重程度上色。"""
    for cell in row:
        cell.alignment = NORMAL_ALIGNMENT
    if severity and severity in SEVERITY_FILLS:
        row[5].fill = SEVERITY_FILLS[severity]


# ─── 报告生成 ───

def write_report(
    output_path: str,
    qc_results: list[dict],
    attribution_results: dict[str, list[dict]],
    stats: dict,
) -> None:
    """
    生成质检报告 Excel，包含三个 sheet。

    Args:
        output_path: 输出文件路径
        qc_results: 合规检测结果列表
        attribution_results: 归因分析结果 {意向结果: [归因条目]}
        stats: 统计概览数据
    """
    wb = openpyxl.Workbook()

    # Sheet 1: 合规检测
    ws_qc = wb.active
    ws_qc.title = "合规检测"
    ws_qc.sheet_properties.tabColor = "4472C4"

    headers = ["id", "时间", "意向结果", "违规规则", "问题类型", "危害程度", "证据片段", "改进建议"]
    for col, h in enumerate(headers, 1):
        _style_header(ws_qc.cell(1, col, h))

    row_num = 2
    for record in qc_results:
        dialog_id = record.get("id", "")
        time_val = record.get("time", "")
        intent = record.get("intent", "")
        violations = record.get("violations", [])

        if not violations:
            # 无违规：一行标记"通过"
            ws_qc.cell(row_num, 1, dialog_id)
            ws_qc.cell(row_num, 2, time_val)
            ws_qc.cell(row_num, 3, intent)
            ws_qc.cell(row_num, 4, "通过")
            row_num += 1
        else:
            for v in violations:
                ws_qc.cell(row_num, 1, dialog_id)
                ws_qc.cell(row_num, 2, time_val)
                ws_qc.cell(row_num, 3, intent)
                ws_qc.cell(row_num, 4, v.get("rule_id", ""))
                ws_qc.cell(row_num, 5, v.get("rule_name", ""))
                ws_qc.cell(row_num, 6, v.get("severity", ""))
                ws_qc.cell(row_num, 7, v.get("evidence", ""))
                ws_qc.cell(row_num, 8, v.get("suggestion", ""))
                _style_row(ws_qc[row_num - 1], v.get("severity"))
                row_num += 1

    # 自适应列宽
    for col in ws_qc.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws_qc.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # Sheet 2: 归因分析
    ws_attr = wb.create_sheet("归因分析")
    ws_attr.sheet_properties.tabColor = "6BCB77"

    attr_headers = ["意向结果", "归因类别", "占比", "数量", "典型案例", "改进建议"]
    for col, h in enumerate(attr_headers, 1):
        _style_header(ws_attr.cell(1, col, h))

    row_num = 2
    for intent, categories in attribution_results.items():
        for cat in categories:
            ws_attr.cell(row_num, 1, intent)
            ws_attr.cell(row_num, 2, cat.get("category", ""))
            ws_attr.cell(row_num, 3, f"{cat.get('ratio', 0) * 100:.1f}%")
            ws_attr.cell(row_num, 4, cat.get("count", 0))
            ws_attr.cell(row_num, 5, "\n".join(cat.get("examples", [])[:3]))
            ws_attr.cell(row_num, 6, cat.get("suggestion", ""))
            row_num += 1

    for col in ws_attr.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws_attr.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # Sheet 3: 统计概览
    ws_stats = wb.create_sheet("统计概览")
    ws_stats.sheet_properties.tabColor = "FFD93D"

    ws_stats.cell(1, 1, "指标")
    ws_stats.cell(1, 2, "数值")
    _style_header(ws_stats.cell(1, 1))
    _style_header(ws_stats.cell(1, 2))

    row_num = 2
    # 基础指标
    for key, label in [("total", "总对话数"), ("pass", "通过数"), ("violation_rate", "违规率")]:
        if key in stats:
            ws_stats.cell(row_num, 1, label)
            val = stats[key]
            ws_stats.cell(row_num, 2, val)
            row_num += 1

    # 空一行分隔
    row_num += 1

    # 规则命中明细：规则 ID | 规则名称 | 命中次数 | 占比
    rule_header = ["规则ID", "规则名称", "命中次数", "占比"]
    for col, h in enumerate(rule_header, 1):
        _style_header(ws_stats.cell(row_num, col, h))
    row_num += 1

    # 从 rules_hit 中提取规则数据
    rules_hit = stats.get("rules_hit", {})
    rule_names = stats.get("rule_names", {})
    # rules_hit 格式：{"R05": 103, "R06": 47, ...}
    total_violations = sum(rules_hit.values()) if rules_hit else 0

    # 按规则 ID 排序输出
    for rule_id in sorted(rules_hit.keys()):
        count = rules_hit[rule_id]
        pct = f"{count / total_violations * 100:.1f}%" if total_violations > 0 else "0.0%"
        ws_stats.cell(row_num, 1, rule_id)
        ws_stats.cell(row_num, 2, rule_names.get(rule_id, ""))
        ws_stats.cell(row_num, 3, count)
        ws_stats.cell(row_num, 4, pct)
        row_num += 1

    # 汇总行
    ws_stats.cell(row_num, 1, "合计")
    ws_stats.cell(row_num, 3, total_violations)
    ws_stats.cell(row_num, 4, "100.0%")
    row_num += 1

    # 统计概览列宽
    for col in ws_stats.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws_stats.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    # 保存
    wb.save(output_path)
    wb.close()


# ─── 临时文件清理 ───

def cleanup_temp_files(work_dir: str, keep_temp: bool = False) -> None:
    """
    清理处理过程中的临时文件。

    Args:
        work_dir: 工作目录（批次 JSON、进度文件所在目录）
        keep_temp: 是否保留临时文件
    """
    if keep_temp:
        print("保留临时文件 (--keep-temp)")
        return

    work = Path(work_dir)

    # 清理进度文件
    for name in ["progress.json", "failed_batches.json"]:
        f = work / name
        if f.exists():
            f.unlink()
            print(f"已删除临时文件: {f}")

    # 清理批次目录
    for name in ["batches", "attribution_batches"]:
        dir_path = work / name
        if dir_path.exists():
            shutil.rmtree(dir_path)
            print(f"已删除批次目录: {dir_path}")

    # 清理中间结果文件
    for f in work.glob("batch_result_*.json"):
        f.unlink()
        print(f"已删除中间结果: {f}")


# ─── CLI 入口 ───

def main():
    parser = argparse.ArgumentParser(description="auto-qc 报告生成器")
    sub = parser.add_subparsers(dest="command", required=True)

    write_p = sub.add_parser("write", help="生成质检报告 Excel")
    write_p.add_argument("--output", required=True, help="报告输出路径")
    write_p.add_argument("--qc-results", required=True, help="合规检测结果 JSON 路径")
    write_p.add_argument("--attribution", help="归因分析结果 JSON 路径")
    write_p.add_argument("--stats", help="统计概览数据 JSON 路径")

    cleanup_p = sub.add_parser("cleanup", help="清理临时文件")
    cleanup_p.add_argument("--dir", required=True, help="工作目录")
    cleanup_p.add_argument("--keep-temp", action="store_true", help="保留临时文件")

    args = parser.parse_args()

    if args.command == "write":
        with open(args.qc_results, "r", encoding="utf-8") as f:
            qc_data = json.load(f)

        attr_data = {}
        if args.attribution:
            if os.path.exists(args.attribution):
                with open(args.attribution, "r", encoding="utf-8") as f:
                    attr_data = json.load(f)
            else:
                print(f"警告：归因分析结果文件不存在: {args.attribution}")

        stats = {}
        if args.stats:
            if os.path.exists(args.stats):
                with open(args.stats, "r", encoding="utf-8") as f:
                    stats = json.load(f)
            else:
                print(f"警告：统计概览数据文件不存在: {args.stats}")

        # 确保输出目录存在
        Path(args.output).parent.mkdir(parents=True, exist_ok=True)

        write_report(args.output, qc_data, attr_data, stats)
        print(f"报告已生成: {args.output}")

    elif args.command == "cleanup":
        cleanup_temp_files(args.dir, args.keep_temp)
        print("临时文件清理完成")


if __name__ == "__main__":
    main()
